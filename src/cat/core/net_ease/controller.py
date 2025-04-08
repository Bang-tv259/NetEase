from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import Workbook, load_workbook

from cat.core.net_ease.dto import SalaryOutput
from cat.core.net_ease.services import (
    calculate_tax,
    calculate_insurance,
    calculate_personal_deduction,
)


if TYPE_CHECKING:
    from fastapi import UploadFile

    from cat.core.net_ease.constants import TaxConfig


def handle_convert_gross_to_net(
    gross_salary: float, number_of_dependents: int, tax_config_dep: TaxConfig
) -> SalaryOutput:
    """Convert gross salary to net salary.

    Apply:
    - personal and dependent deductions
    - insurance contributions
    - personal income tax
    """
    # Step 1: Calculate insurance once
    insurance_amount = calculate_insurance(gross_salary)

    # Step 2: Compute taxable income (pre-tax income)
    pre_tax_income = (
        gross_salary
        - insurance_amount
        - calculate_personal_deduction(number_of_dependents)
    )

    # Step 3: Calculate personal income tax
    personal_income_tax = calculate_tax(pre_tax_income, tax_config_dep)

    # Step 4: Compute net salary
    net_salary = gross_salary - insurance_amount - personal_income_tax

    return SalaryOutput(
        gross_salary=gross_salary,
        net_salary=net_salary,
        insurance_amount=insurance_amount,
        personal_income_tax=personal_income_tax,
    )


async def handle_upload_excel(file: UploadFile, tax_config_dep: TaxConfig) -> Workbook:
    # Read the uploaded file content
    file_content = await file.read()

    # Load the workbook and select the active sheet
    wb = load_workbook(filename=BytesIO(file_content))
    sheet = wb.active

    # Prepare a new workbook to write the result to
    new_wb = Workbook()
    new_sheet = new_wb.active
    new_sheet.append([
        "ID",
        "Employee Name",
        "Gross Salary",
        "Number of Dependents",
        "Net Salary",
    ])

    # Iterate through the rows in the original sheet
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the row is empty
        if all(cell is None for cell in row):
            break

        # Extract data from the row
        employee_id = row[0]
        employee_name = row[1]
        gross_salary = row[2]
        number_of_dependents = row[3]

        # Call the conversion function
        net_salary_output = handle_convert_gross_to_net(
            float(gross_salary), int(number_of_dependents), tax_config_dep
        )

        # Write the result to the new sheet
        new_sheet.append([
            employee_id,
            employee_name,
            gross_salary,
            number_of_dependents,
            net_salary_output.net_salary,
        ])

    return new_wb
